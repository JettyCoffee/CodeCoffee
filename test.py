# Version:      2.1
# StartTime:   2021/1/6 12:30
# Finished:    2021/1/7 20:30
# Author:      MickLife
# B站:         https://space.bilibili.com/38626658
 
# 补充内容大致如下：
# 修改了代码中支付宝和微信csv文件列名
# 支付宝csv文件使用gbk编码读取
# 删除了文字小尾巴使其在PowerBi上更具有可读性
# 增加了一些辅助性质的打印信息
# 修改时间：2023/10/16
 
 
import pandas as pd
import openpyxl
import tkinter.filedialog
import msvcrt
 
#设置显示的最大列、宽等参数，消掉打印不完全中间的省略号
pd.set_option('display.max_columns', 1000)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 1000)
 
def strip_in_data(data):  # 把列名中和数据中首尾的空格都去掉。
    data = data.rename(columns={column_name: column_name.strip() for column_name in data.columns})
    data = data.apply(lambda x: x.map(lambda y: y.strip().strip('¥') if isinstance(y, str) else y))
    return data
 
def read_data_wx():  # 获取微信数据 #（path）
    d_wx = pd.read_csv(path_wx, header=16, skipfooter=0, encoding='utf-8')  # 数据获取，微信 #把path改为path_wx了,-sig
    d_wx = d_wx.iloc[:, [0, 4, 7, 1, 2, 3, 5]]  # 按顺序提取所需列
    d_wx = strip_in_data(d_wx)  # 去除列名与数值中的空格。
    a = d_wx.iloc[:, 0]
    b = pd.to_datetime(a)
    date_only = b.dt.date
    d_wx.iloc[:, 0] = date_only  # 日期类型更改
    c = d_wx.iloc[:, 6]
    d = c.astype('float64')
    d_wx.iloc[:, 6] = d  # 浮点数类型更改
    d_wx = d_wx.drop(d_wx[d_wx['收/支'] == '/'].index)  # 删除'收/支'为'/'的行
    d_wx.rename(columns={'当前状态': '支付状态', '交易类型': '类型', '金额(元)': '金额'}, inplace=True)  # 修改列名称
    d_wx.insert(1, '来源', "微信", allow_duplicates=True)  # 添加微信来源标识
    len1 = len(d_wx)
    print("成功读取 " + str(len1) + " 条「微信」账单数据\n")
    return d_wx
 
def read_data_zfb():  # 获取支付宝数据
    # 读取ANSI编码的CSV文件
    d_zfb = pd.read_csv(path_zfb, header=4, skipfooter=7, encoding='gbk', engine='python')
    d_zfb = d_zfb.iloc[:, [2, 10, 11, 6, 7, 8, 9]]  # 按顺序提取所需列
    d_zfb = strip_in_data(d_zfb)  # 去除列名与数值中的空格
    g = d_zfb.iloc[:, 0]
    b = pd.to_datetime(g)
    date_only = b.dt.date
    d_zfb.iloc[:, 0] = date_only    # 修改时间类型
    c = d_zfb.iloc[:, 6]
    d = c.astype('float64')
    d_zfb.iloc[:, 6] = d  # 修改浮点数类型
    d_zfb = d_zfb.drop(d_zfb[d_zfb['收/支'] == '不计收支'].index)  # 删除'收/支'为'不计收支'的行
    d_zfb.rename(columns={'交易创建时间': '交易时间', '商品名称': '商品', '金额（元）': '金额', '交易状态': '支付状态'}, inplace=True)  # 修改列名称
    d_zfb.insert(1, '来源', "支付宝", allow_duplicates=True)  # 添加支付宝来源标识
    len1 = len(d_zfb)
    print("成功读取 " + str(len1) + " 条「支付宝」账单数据\n")
    return d_zfb
 
def add_cols(data):  # 增加3列数据
    # 逻辑1：取值-1 or 1。-1表示支出，1表示收入。
    data.insert(8, '逻辑1', -1, allow_duplicates=True)  # 插入列，默认值为-1
    for index in range(len(data.iloc[:, 2])):  # 遍历第3列的值，判断为收入，则改'逻辑1'为1
        if data.iloc[index, 2] == '收入':
            data.iloc[index, 8] = 1
 
    # 逻辑2：取值0 or 1。1表示计入，0表示不计入。
    data.insert(9, '逻辑2', 1, allow_duplicates=True)  # 插入到第10列，默认值为1
    for index in range(len(data.iloc[:, 3])):  # 遍历第4列的值，判断为资金流动，则改'逻辑2'为0
        if data.iloc[index, 3] == '交易成功':
            data.iloc[index, 9] = 1
    # 月份
    data.insert(1, '月份', 0, allow_duplicates=True)  # 插入列，第二列，默认值为0
    for index in range(len(data.iloc[:, 0])):
        time = data.iloc[index, 0]
        data.iloc[index, 1] = time.month  # 访问月份属性的值，赋给这月份列
 
    # 乘后金额
    data.insert(11, '乘后金额', 0, allow_duplicates=True)  # 插入列，默认值为0
    for index in range(len(data.iloc[:, 8])):  # 遍历逻辑1
        money = data.iloc[index, 8] * data.iloc[index, 9] * data.iloc[index, 10]
        data.iloc[:, 11] = data.iloc[:, 11].astype(float)
        data.iloc[index, 11] = money  # 插入乘后金额到12列
    return data
# 从这里开始执行代码
if __name__ == '__main__':
 
    # 路径设置
    print('提示：请在弹窗中选择要导入的【微信】账单文件\n')
    path_wx = tkinter.filedialog.askopenfilename(title='选择要导入的微信账单：', filetypes=[('所有文件', '.*'), ('csv文件', '.csv')])
    if path_wx == '':  # 判断是否只导入了微信或支付宝账单中的一个
        cancel_wx = 1
        print('wx导入失败')   # 本人
    else:
        cancel_wx = 0
        print('wx导入成功')  # 本人
 
    print('提示：请在弹窗中选择要导入的【支付宝】账单文件\n')
    path_zfb = tkinter.filedialog.askopenfilename(title='选择要导入的支付宝账单：', filetypes=[('所有文件', '.*'), ('csv文件', '.csv')])
    if path_zfb == '':  # 判断是否只导入了微信或支付宝账单中的一个
        cancel_zfb = 1
        print('zfb导入失败')  # 本人
    else:
        cancel_zfb = 0
        print('zfb导入成功')  # 本人
 
    while cancel_zfb == 1 and cancel_wx == 1:
        print('\n您没有选择任何一个账单！     请按任意键退出程序')
        ord(msvcrt.getch())
 
    path_account = tkinter.filedialog.askopenfilename(title='选择要导出的目标账本表格：', filetypes=[('所有文件', '.*'), ('Excel表格', '.xlsx')])
    while path_account == '':  # 判断是否选择了账本
        print('\n年轻人，不选账本怎么记账？      请按任意键退出程序')
        ord(msvcrt.getch())
        # path空值同样失败
        # path_account 是合并输出数据
 
    path_write = path_account
 
    # 判断是否只导入了微信或支付宝账单中的一个
    if cancel_wx == 1:
        data_wx = pd.DataFrame()
    else:
        data_wx = read_data_wx()
    if cancel_zfb == 1:
        data_zfb = pd.DataFrame()
    else:
        data_zfb = read_data_zfb()
 
        type_c = 'data_wx的类型：' + str(type(data_wx))
        print(type_c)
 
        print('******************************')
 
        type_d = 'data_zfb的类型：' + str(type(data_zfb))
        print(type_c)
 
    data_merge = pd.concat([data_wx, data_zfb], axis=0)  # 上下拼接合并表格
    print(data_merge)
    data_merge = add_cols(data_merge)  # 新增 逻辑、月份、乘后金额 3列
    print("已自动计算乘后金额和交易月份，已合并数据")
    merge_list = data_merge.values.tolist()  # 格式转换，DataFrame->List
    workbook = openpyxl.load_workbook(path_account)  # openpyxl读取账本文件
 
    sheet = workbook['明细']
    max_data_row = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):  # 从第二行开始遍历
        if all(cell.value is None or cell.value == '' for cell in row):
            break
        max_data_row = row[0].row
 
    print('\n「明细」 sheet 页已有 ' + str(max_data_row) + ' 行数据，将在末尾写入数据')
 
    for row in merge_list:
        sheet.append(row)  # 将每一行数据追加到工作表中
 
    workbook.save(path_write)  # 保存
    print("\n成功将数据写入到 " + path_write)
    print("\n运行成功！write successfully!    按任意键退出")
    ord(msvcrt.getch())
 
 